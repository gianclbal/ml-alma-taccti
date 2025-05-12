from fastapi import FastAPI, Depends, HTTPException, File, UploadFile, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List
import re
import pandas as pd
import uuid
import os
from datetime import datetime
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from app.inference import predict_sentences, load_model_for_theme
from app.auth import router as auth_router, get_current_user
from openpyxl.styles import Alignment

app = FastAPI()
app.include_router(auth_router, prefix="/auth")


class EssayInput(BaseModel):
    id: str
    essay: str

class AnalyzeRequest(BaseModel):
    thematic_code: int
    essays: List[EssayInput]

# Add CORS middleware to allow requests from localhost:3000 (your React app)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Allows your React app to make requests to the backend
    allow_credentials=True,
    allow_methods=["*"],  # Allows all HTTP methods (GET, POST, etc.)
    allow_headers=["*"],  # Allows all headers
)


# Serve files from the "static" directory
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.post("/analyze")
def analyze_essays(payload: AnalyzeRequest, current_user: str = Depends(get_current_user)):
    response_data = []
    theme_present = 0
    theme_absent = 0

    # Set default thematic_code to 1 (Aspirational) if it's not provided or invalid
    thematic_code = payload.thematic_code if payload.thematic_code in THEME_CONFIG else 1

    for item in payload.essays:
        essay_id = item.id
        original_essay = item.essay
        sentence_list = re.split(r'[.?!]', original_essay)
        sentence_list = [s.strip() for s in sentence_list if s.strip()]

        # Get the model path from the thematic_code
        model_path = THEME_CONFIG[thematic_code]["model_path"]

        # Call the prediction function with the correct model path
        y_pred_results = predict_sentences(sentence_list, model_path)
        predictions = [r["prediction"] for r in y_pred_results]

        if max(predictions) > 0:
            theme_present += 1
            marked_essay = original_essay
            for i, result in enumerate(y_pred_results):
                if result["prediction"] == 1:
                    marked_essay = marked_essay.replace(sentence_list[i], f"<mark>{sentence_list[i]}</mark>")
            theme_label = "Yes"
        else:
            theme_absent += 1
            marked_essay = original_essay
            theme_label = "No"

        response_data.append({
            "Essay ID": essay_id,
            "Year": "",
            "Semester": "",
            "Class": "",
            "Type": "",
            "Section": "",
            "Alma ID": "",
            "Aspirational Present": theme_label,
            "Annotated Essays": marked_essay
        })

    return {
        "results": response_data,
        "chart_counts": [theme_present, theme_absent]
    }

THEME_CONFIG = {    1: {"name": "Aspirational", "model_path": "gianclbal/asp_deberta", "color": "FF0000"},
    2: {"name": "Familial", "model_path": "gianclbal/fam_deberta", "color": "800080"},
    3: {"name": "Navigational", "model_path": "gianclbal/nav_deberta", "color": "006400"},
    4: {"name": "Resistance", "model_path": "gianclbal/res_deberta", "color": "FF8C00"},
    5: {"name": "Social", "model_path": "gianclbal/soc_deberta", "color": "0000FF"}}

LOCAL_THEME_CONFIG = {    1: {"name": "Aspirational", "model_path": "gianclbal/asp_deberta", "color": "FF0000"},
    2: {"name": "Familial", "model_path": "/Users/gbaldonado/Developer/ml-alma-taccti/ml-alma-taccti/webapp/models/fam_plus_deberta_batch_1_2_runtime_model", "color": "800080"},
    3: {"name": "Navigational", "model_path": "/Users/gbaldonado/Developer/ml-alma-taccti/ml-alma-taccti/webapp/models/nav_plus_deberta_batch_1_2_runtime_model", "color": "006400"},
    4: {"name": "Resistance", "model_path": "/Users/gbaldonado/Developer/ml-alma-taccti/ml-alma-taccti/webapp/models/res_plus_deberta_batch_1_2_runtime_model", "color": "FF8C00"},
    5: {"name": "Social", "model_path": "/Users/gbaldonado/Developer/ml-alma-taccti/ml-alma-taccti/webapp/models/soc_plus_deberta_batch_1_2_runtime_model", "color": "0000FF"}}

# Function to analyze essays based on a given thematic code
@app.post("/analyze-file")
def analyze_file(
    file: UploadFile = File(...),
    thematic_code: int = Form(...),
    id_column: str = Form(...),
    essay_column: str = Form(...),
    current_user: str = Depends(get_current_user)
):
    # Ensure the thematic_code is valid
    if thematic_code not in THEME_CONFIG:
        raise HTTPException(status_code=400, detail=f"Invalid thematic code. Available codes are {list(THEME_CONFIG.keys())}.")

    # Load the model using the thematic_code to get the model path
    model_config = THEME_CONFIG[thematic_code]
    model_path = model_config["model_path"]
    theme_name = model_config["name"]

    # Load the model
    predictor = load_model_for_theme(model_path)
    if predictor is None:
        raise HTTPException(status_code=500, detail="Model loading failed.")

    # Read the CSV file
    df = pd.read_csv(file.file)
    if id_column not in df.columns or essay_column not in df.columns:
        raise HTTPException(status_code=400, detail=f"CSV must contain '{id_column}' and '{essay_column}' columns.")

    results = []
    theme_present = 0
    theme_absent = 0

    for _, row in df.iterrows():
        essay_id = row[id_column]
        original_essay = row[essay_column]
        sentence_list = re.split(r'[.?!]', original_essay)
        sentence_list = [s.strip() for s in sentence_list if s.strip()]

        # Get model predictions for each sentence
        y_pred_results = predictor(sentence_list)
        predictions = [r["prediction"] for r in y_pred_results]

        if max(predictions) > 0:
            theme_present += 1
            marked_essay = original_essay
            for i, result in enumerate(y_pred_results):
                if result["prediction"] == 1:
                    marked_essay = marked_essay.replace(sentence_list[i], f"<mark>{sentence_list[i]}</mark>")
            theme_label = "Yes"
        else:
            theme_absent += 1
            marked_essay = original_essay
            theme_label = "No"

        results.append({
            "Essay ID": essay_id,
            "Year": "",
            "Semester": "",
            "Class": "",
            "Type": "",
            "Section": "",
            "Alma ID": "",
            f"{theme_name} Present": theme_label,
            "Annotated Essays": marked_essay
        })

    # Prepare the output file
    # Get the current date
    current_date = datetime.now().strftime("%Y-%m-%d")
    out_df = pd.DataFrame(results)
    output_filename = f"{theme_name}_{current_date}_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join("static", output_filename)  # Move the file to a publicly accessible folder

    
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="Theme_1")

        workbook = writer.book
        worksheet = writer.sheets["Theme_1"]

        red_font = InlineFont(color="FF0000")
        red_font.b = True

        for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, min_col=worksheet.max_column):
            for cell in row:
                if cell.value and "<mark>" in str(cell.value):
                    parts = re.split(r"(<mark>.*?</mark>)", str(cell.value))
                    rt = CellRichText()
                    for part in parts:
                        if part.startswith("<mark>") and part.endswith("</mark>"):
                            text = part.replace("<mark>", "").replace("</mark>", "")
                            rt.append(TextBlock(text=text, font=red_font))
                        else:
                            rt.append(TextBlock(text=part, font=InlineFont()))
                    cell.value = rt

        column_index = 9  # Column I (1-based index)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=column_index, max_col=column_index):
            for cell in row:
                # Set text wrapping
                cell.alignment = Alignment(wrap_text=True)
        
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 50  # Adjust width as needed

    # Return a URL for the frontend to download the file
    file_url = f"http://127.0.0.1:8001/static/{output_filename}"  # Serve from the static folder

    return JSONResponse(
        content={
            "file_url": file_url,
            "results": results,  # Your results data
            "filename": output_filename,
            "theme_name": theme_name,
        },
        status_code=200
    )

def process_essays(df, predictor, id_column, essay_column, theme_name):
    results = []
    theme_present = 0
    theme_absent = 0

    for _, row in df.iterrows():
        essay_id = row[id_column]
        original_essay = row[essay_column]
        sentence_list = re.split(r'[.?!]', original_essay)  # Split the essay into sentences
        sentence_list = [s.strip() for s in sentence_list if s.strip()]

        # Get model predictions for each sentence
        y_pred_results = predictor(sentence_list)
        predictions = [r["prediction"] for r in y_pred_results]

        if max(predictions) > 0:
            theme_present += 1
            marked_essay = original_essay
            for i, result in enumerate(y_pred_results):
                if result["prediction"] == 1:
                    marked_essay = marked_essay.replace(sentence_list[i], f"<mark>{sentence_list[i]}</mark>")
            theme_label = "Yes"
        else:
            theme_absent += 1
            marked_essay = original_essay
            theme_label = "No"

        # Collect results for the current essay
        results.append({
            "Essay ID": essay_id,
            "Year": "",
            "Semester": "",
            "Class": "",
            "Type": "",
            "Section": "",
            "Alma ID": "",
            f"{theme_name} Present": theme_label,
            "Annotated Essays": marked_essay
        })

    return results

@app.post("/analyze-all-themes")
def analyze_all_themes(
    file: UploadFile = File(...),
    id_column: str = Form(...),
    essay_column: str = Form(...),
    current_user: str = Depends(get_current_user)
):
    # Check if file is CSV or XLSX
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Please upload a valid CSV or XLSX file.")

    # Read the file content based on its extension
    if file.filename.endswith('.csv'):
        df = pd.read_csv(file.file)
    elif file.filename.endswith('.xlsx'):
        df = pd.read_excel(file.file)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format.")

    # Check for necessary columns in the uploaded file
    if id_column not in df.columns or essay_column not in df.columns:
        raise HTTPException(status_code=400, detail=f"CSV/XLSX must contain '{id_column}' and '{essay_column}' columns.")

    # Create the output file path and name
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_filename = f"All_Themes_{current_date}_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join("static", output_filename)

    results = [] # store the results for all themes in json format

    # Prepare the Excel writer
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Iterate over themes, process essays, and save as separate sheets
        for theme_code, theme_config in THEME_CONFIG.items():
            theme_name = theme_config["name"]
            model_path = theme_config["model_path"]

            # Load the model for the specific theme
            predictor = load_model_for_theme(model_path)
            if predictor is None:
                raise HTTPException(status_code=500, detail=f"Failed to load model for {theme_name}")

            # Process essays for the theme
            theme_results = process_essays(df, predictor, id_column, essay_column, theme_name)
            
            # Log to console that essays for this theme are processed
            print(f"Processed essays for theme: {theme_name}")

            # Convert theme results to DataFrame
            theme_df = pd.DataFrame(theme_results)

            # Save to an Excel sheet
            theme_df.to_excel(writer, index=False, sheet_name=theme_name)

            # Get the workbook and worksheet for rich text manipulation
            workbook = writer.book
            worksheet = writer.sheets[theme_name]

            # Apply rich text formatting for cells containing <mark> tags
            red_font = InlineFont(color="FF0000")
            red_font.b = True

            for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, min_col=worksheet.max_column):
                for cell in row:
                    if cell.value and "<mark>" in str(cell.value):
                        parts = re.split(r"(<mark>.*?</mark>)", str(cell.value))
                        rt = CellRichText()
                        for part in parts:
                            if part.startswith("<mark>") and part.endswith("</mark>"):
                                text = part.replace("<mark>", "").replace("</mark>", "")
                                rt.append(TextBlock(text=text, font=red_font))
                            else:
                                rt.append(TextBlock(text=part, font=InlineFont()))
                        cell.value = rt
            
            # Apply alignment and width to the "Annotated Essays" column (Column I)
            column_index = 9  # Column I (1-based index)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=column_index, max_col=column_index):
                for cell in row:
                    # Set text wrapping
                    cell.alignment = Alignment(wrap_text=True)
            
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 50  # Adjust width as needed

             # Add the theme results to the final results list in JSON format
            results.append({
                "theme_name": theme_name,
                "results": theme_results  # Store results for this theme in JSON
            })

    # Return the file URL for the frontend to download the file
    file_url = f"http://127.0.0.1:8001/static/{output_filename}"

    return JSONResponse(
        content={
            "file_url": file_url,
            "filename": output_filename,
            "results": results,  # Include the results for all themes in JSON format
        },
        status_code=200
    )


# @app.post("/analyze-aspirational")
# def analyze_file(
#     file: UploadFile = File(...),
#     thematic_code: int = Form(...),
#     id_column: str = Form(...),
#     essay_column: str = Form(...),
#     current_user: str = Depends(get_current_user)
# ):
#     df = pd.read_csv(file.file)
#     if id_column not in df.columns or essay_column not in df.columns:
#         raise HTTPException(status_code=400, detail=f"CSV must contain '{id_column}' and '{essay_column}' columns.")

#     results = []
#     theme_present = 0
#     theme_absent = 0

#     for _, row in df.iterrows():
#         essay_id = row[id_column]
#         original_essay = row[essay_column]
#         sentence_list = re.split(r'[.?!]', original_essay)
#         sentence_list = [s.strip() for s in sentence_list if s.strip()]

#         y_pred_results = predict_sentences(sentence_list)
#         predictions = [r["prediction"] for r in y_pred_results]

#         if max(predictions) > 0:
#             theme_present += 1
#             marked_essay = original_essay
#             for i, result in enumerate(y_pred_results):
#                 if result["prediction"] == 1:
#                     marked_essay = marked_essay.replace(sentence_list[i], f"<mark>{sentence_list[i]}</mark>")
#             theme_label = "Yes"
#         else:
#             theme_absent += 1
#             marked_essay = original_essay
#             theme_label = "No"

#         results.append({
#             "Essay ID": essay_id,
#             "Year": "",
#             "Semester": "",
#             "Class": "",
#             "Type": "",
#             "Section": "",
#             "Alma ID": "",
#             "Aspirational Present": theme_label,
#             "Annotated Essays": marked_essay
#         })

#     out_df = pd.DataFrame(results)
#     output_filename = f"analyzed_{uuid.uuid4().hex[:8]}.xlsx"
#     output_path = os.path.join("/tmp", output_filename)
    
#     with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
#         out_df.to_excel(writer, index=False, sheet_name="Theme_1")

#         workbook = writer.book
#         worksheet = writer.sheets["Theme_1"]

#         red_font = InlineFont(color="FF0000")
#         red_font.b = True

#         for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, min_col=worksheet.max_column):
#             for cell in row:
#                 if cell.value and "<mark>" in str(cell.value):
#                     parts = re.split(r"(<mark>.*?</mark>)", str(cell.value))
#                     rt = CellRichText()
#                     for part in parts:
#                         if part.startswith("<mark>") and part.endswith("</mark>"):
#                             text = part.replace("<mark>", "").replace("</mark>", "")
#                             rt.append(TextBlock(text=text, font=red_font))
#                         else:
#                             rt.append(TextBlock(text=part, font=InlineFont()))
#                     cell.value = rt

#     return FileResponse(path=output_path, filename=output_filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")